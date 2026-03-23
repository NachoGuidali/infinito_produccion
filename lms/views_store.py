"""
lms/views_store.py
Vistas de la tienda de productos físicos.
Se mantiene separado de views.py para no tocarlo.
"""
import io
import time
from datetime import datetime, timezone as dt_timezone
from decimal import Decimal

import openpyxl
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.files.storage import default_storage
from django.core.paginator import Paginator
from django.db import transaction
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils.text import slugify
from django.views.decorators.http import require_POST, require_GET

from .models import (
    Product,
    ProductCategory,
    ShippingConfig,
    ShippingZone,
    StoreOrder,
    StoreOrderItem,
)
from .services.payments import create_mp_preference_for_store_order

TAB_ITEMS = [
    ("products",   "📦 Productos"),
    ("categories", "🏷️ Categorías"),
    ("orders",     "🚚 Pedidos"),
    ("shipping",   "✉️ Envíos"),
]

# ─── helpers ────────────────────────────────────────────────────────────────

def _is_staff(u):
    return u.is_authenticated and u.is_staff


def _store_cart(session):
    c = session.get("store_cart")
    if not isinstance(c, dict):
        c = {}
        session["store_cart"] = c
    return c


def _cart_totals(cart: dict):
    total = Decimal("0")
    for item in cart.values():
        try:
            total += Decimal(item["price_ars"]) * int(item["qty"])
        except Exception:
            pass
    return total


def _cart_items_for_shipping(cart: dict):
    result = []
    for item in cart.values():
        try:
            p = Product.objects.get(id=item["product_id"])
            result.append({"product": p, "qty": int(item["qty"])})
        except Product.DoesNotExist:
            pass
    return result


# ─── TIENDA (catálogo público) ────────────────────────────────────────────────

def store(request):
    qs = Product.objects.filter(is_active=True).select_related("category")
    category_slug = request.GET.get("cat")
    q = request.GET.get("q", "").strip()

    if category_slug:
        qs = qs.filter(category__slug=category_slug)
    if q:
        qs = qs.filter(name__icontains=q)

    paginator = Paginator(qs, 12)
    page = paginator.get_page(request.GET.get("page", 1))

    categories = ProductCategory.objects.all()
    cart = _store_cart(request.session)
    cart_count = sum(int(i["qty"]) for i in cart.values())

    return render(request, "lms/store/store_tienda.html", {
        "page_obj":     page,
        "categories":   categories,
        "selected_cat": category_slug or "",
        "search_q":     q,
        "cart_count":   cart_count,
    })


def store_product_detail(request, slug):
    product = get_object_or_404(Product, slug=slug, is_active=True)
    related = Product.objects.filter(
        category=product.category, is_active=True
    ).exclude(id=product.id)[:4]

    cart = _store_cart(request.session)
    cart_count = sum(int(i["qty"]) for i in cart.values())

    return render(request, "lms/store/product_detail_tienda.html", {
        "product":    product,
        "related":    related,
        "cart_count": cart_count,
    })


# ─── CARRITO ──────────────────────────────────────────────────────────────────

def store_cart(request):
    cart = _store_cart(request.session)
    items = []
    for key, item in cart.items():
        try:
            p = Product.objects.get(id=item["product_id"])
            items.append({
                "key":      key,
                "product":  p,
                "qty":      int(item["qty"]),
                "subtotal": p.price_ars * int(item["qty"]),
            })
        except Product.DoesNotExist:
            pass

    total = sum(i["subtotal"] for i in items)
    return render(request, "lms/store/cart_tienda.html", {
        "items":      items,
        "total":      total,
        "cart_count": sum(i["qty"] for i in items),
    })


@require_POST
def store_cart_add(request, product_id):
    product = get_object_or_404(Product, id=product_id, is_active=True)
    if product.stock <= 0:
        messages.error(request, "Este producto no tiene stock disponible.")
        return redirect(request.META.get("HTTP_REFERER", reverse("lms:store")))

    cart = _store_cart(request.session)
    key = str(product_id)
    if key in cart:
        new_qty = min(int(cart[key]["qty"]) + 1, product.stock)
        cart[key]["qty"] = new_qty
    else:
        cart[key] = {
            "product_id": product.id,
            "name":       product.name,
            "price_ars":  str(product.price_ars),
            "qty":        1,
        }
    request.session.modified = True
    messages.success(request, f'"{product.name}" agregado al carrito.')

    next_url = request.POST.get("next") or request.META.get("HTTP_REFERER", reverse("lms:store"))
    return redirect(next_url)


@require_POST
def store_cart_remove(request, key):
    cart = _store_cart(request.session)
    cart.pop(key, None)
    request.session.modified = True
    messages.success(request, "Producto quitado del carrito.")
    return redirect(reverse("lms:store_cart"))


@require_POST
def store_cart_update(request, key):
    cart = _store_cart(request.session)
    if key in cart:
        try:
            qty     = int(request.POST.get("qty", 1))
            product = Product.objects.get(id=int(cart[key]["product_id"]))
            cart[key]["qty"] = max(1, min(qty, product.stock))
            request.session.modified = True
        except Exception:
            pass
    return redirect(reverse("lms:store_cart"))


@require_GET
def store_shipping_calc(request):
    postal_code = request.GET.get("postal_code", "").strip()
    if not postal_code:
        return JsonResponse({"error": "Código postal requerido"}, status=400)

    cart = _store_cart(request.session)
    if not cart:
        return JsonResponse({"error": "Carrito vacío"}, status=400)

    cart_items = _cart_items_for_shipping(cart)
    config     = ShippingConfig.get()
    cost       = config.calc_shipping(cart_items, postal_code)

    zone = None
    for z in ShippingZone.objects.all():
        if postal_code in z.codes_list():
            zone = z
            break
    if zone is None:
        zone = ShippingZone.objects.order_by("-multiplier").first()

    subtotal = float(_cart_totals(cart))
    return JsonResponse({
        "postal_code":   postal_code,
        "zone_name":     zone.name if zone else "Interior",
        "shipping_cost": cost,
        "free_shipping": cost == 0,
        "free_over":     float(config.free_over_ars),
        "subtotal":      subtotal,
        "total":         subtotal + cost,
    })


# ─── CHECKOUT ────────────────────────────────────────────────────────────────

@login_required
def store_checkout(request):
    cart = _store_cart(request.session)
    if not cart:
        messages.warning(request, "Tu carrito está vacío.")
        return redirect("lms:store")

    cart_items = _cart_items_for_shipping(cart)
    config     = ShippingConfig.get()
    prof       = getattr(request.user, "profile", None)

    if request.method == "POST":
        required = ["recipient_name", "email", "address", "city", "province", "postal_code", "payment_method"]
        data     = {k: request.POST.get(k, "").strip() for k in required + ["phone", "notes"]}
        missing  = [k for k in required if not data[k]]

        if missing:
            messages.error(request, "Completá todos los campos obligatorios.")
        else:
            postal_code   = data["postal_code"]
            shipping_cost = config.calc_shipping(cart_items, postal_code)
            subtotal      = float(_cart_totals(cart))
            total         = subtotal + shipping_cost

            with transaction.atomic():
                order = StoreOrder.objects.create(
                    user           = request.user,
                    status         = "pending",
                    payment_method = data["payment_method"],
                    recipient_name = data["recipient_name"],
                    email          = data["email"],
                    phone          = data["phone"],
                    address        = data["address"],
                    city           = data["city"],
                    province       = data["province"],
                    postal_code    = postal_code,
                    subtotal_ars   = Decimal(str(subtotal)),
                    shipping_ars   = Decimal(str(shipping_cost)),
                    total_ars      = Decimal(str(total)),
                    notes          = data["notes"],
                )
                for item in cart_items:
                    StoreOrderItem.objects.create(
                        order      = order,
                        product    = item["product"],
                        qty        = item["qty"],
                        unit_price = item["product"].price_ars,
                    )
                    item["product"].stock = max(0, item["product"].stock - item["qty"])
                    item["product"].save(update_fields=["stock"])

            receipt = request.FILES.get("receipt")
            if receipt and data["payment_method"] == "transfer":
                safe  = f"store_receipts/order_{order.id}_{int(time.time())}_{slugify(receipt.name)}"
                saved = default_storage.save(safe, receipt)
                order.transfer_receipt = saved
                order.save(update_fields=["transfer_receipt"])

            request.session["store_cart"] = {}
            request.session.modified = True

            if data["payment_method"] == "mp":
                _, init_point = create_mp_preference_for_store_order(order)
                if init_point:
                    return redirect(init_point)

            return redirect(reverse("lms:store_order_success", args=[order.id]))

    pre_postal   = getattr(prof, "postal_code", "") or ""
    pre_shipping = config.calc_shipping(cart_items, pre_postal) if pre_postal else None

    items_display = [
        {
            "product":  item["product"],
            "qty":      item["qty"],
            "subtotal": item["product"].price_ars * item["qty"],
        }
        for item in cart_items
    ]

    return render(request, "lms/store/checkout.html", {
        "items":        items_display,
        "subtotal":     _cart_totals(cart),
        "pre_shipping": pre_shipping,
        "pre_postal":   pre_postal,
        "config":       config,
        "profile":      prof,
        "user":         request.user,
    })


@login_required
def store_order_success(request, order_id):
    order = get_object_or_404(StoreOrder, id=order_id, user=request.user)
    return render(request, "lms/store/order_success_tienda.html", {"order": order})


@login_required
def store_my_orders(request):
    orders = StoreOrder.objects.filter(user=request.user).prefetch_related("items__product")
    return render(request, "lms/store/my_orders_tienda.html", {"orders": orders})


# ─── ADMIN TIENDA ─────────────────────────────────────────────────────────────

@user_passes_test(_is_staff)
def store_admin(request):
    tab           = request.GET.get("tab", "products")
    status_filter = request.GET.get("order_status", "")

    products   = Product.objects.select_related("category").order_by("-created_at")
    categories = ProductCategory.objects.all()
    config     = ShippingConfig.get()
    zones      = ShippingZone.objects.all()

    orders_qs = StoreOrder.objects.select_related("user").prefetch_related("items__product").order_by("-created_at")
    if status_filter:
        orders_qs = orders_qs.filter(status=status_filter)
    orders = orders_qs[:50]

    return render(request, "lms/store/admin_dashboard_tienda.html", {
        "products":       products,
        "categories":     categories,
        "orders":         orders,
        "config":         config,
        "zones":          zones,
        "tab":            tab,
        "tab_items":      TAB_ITEMS,
        "status_filter":  status_filter,
        "order_statuses": StoreOrder.STATUS,
    })


# ── PRODUCTOS ──

@user_passes_test(_is_staff)
def store_admin_product_create(request):
    categories = ProductCategory.objects.all()
    if request.method == "POST":
        d   = request.POST
        img = request.FILES.get("image")
        p   = Product(
            name        = d.get("name", "").strip(),
            slug        = slugify(d.get("name", "")) + f"-{int(time.time())}",
            description = d.get("description", "").strip(),
            price_ars   = Decimal(d.get("price_ars") or "0"),
            stock       = int(d.get("stock") or 0),
            image_url   = d.get("image_url", "").strip(),
            weight_kg   = Decimal(d.get("weight_kg") or "0"),
            width_cm    = Decimal(d.get("width_cm") or "0"),
            height_cm   = Decimal(d.get("height_cm") or "0"),
            depth_cm    = Decimal(d.get("depth_cm") or "0"),
            is_active   = d.get("is_active") == "true",
        )
        cat_id = d.get("category")
        if cat_id:
            p.category_id = int(cat_id)
        if img:
            p.image = img
        p.save()
        messages.success(request, f'Producto "{p.name}" creado.')
        return redirect(f"{reverse('lms:store_admin')}?tab=products")

    return render(request, "lms/store/admin_product_form_tienda.html", {
        "categories": categories, "product": None
    })


@user_passes_test(_is_staff)
def store_admin_product_edit(request, product_id):
    product    = get_object_or_404(Product, id=product_id)
    categories = ProductCategory.objects.all()

    if request.method == "POST":
        d   = request.POST
        img = request.FILES.get("image")
        product.name        = d.get("name", "").strip()
        product.description = d.get("description", "").strip()
        product.price_ars   = Decimal(d.get("price_ars") or "0")
        product.stock       = int(d.get("stock") or 0)
        product.image_url   = d.get("image_url", "").strip()
        product.weight_kg   = Decimal(d.get("weight_kg") or "0")
        product.width_cm    = Decimal(d.get("width_cm") or "0")
        product.height_cm   = Decimal(d.get("height_cm") or "0")
        product.depth_cm    = Decimal(d.get("depth_cm") or "0")
        product.is_active   = d.get("is_active") == "true"
        cat_id = d.get("category")
        product.category_id = int(cat_id) if cat_id else None
        if img:
            product.image = img
        product.save()
        messages.success(request, f'Producto "{product.name}" actualizado.')
        return redirect(f"{reverse('lms:store_admin')}?tab=products")

    return render(request, "lms/store/admin_product_form_tienda.html", {
        "categories": categories, "product": product
    })


@user_passes_test(_is_staff)
@require_POST
def store_admin_product_delete(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    name    = product.name
    product.delete()
    messages.success(request, f'Producto "{name}" eliminado.')
    return redirect(f"{reverse('lms:store_admin')}?tab=products")


@user_passes_test(_is_staff)
@require_POST
def store_admin_product_import(request):
    from decimal import InvalidOperation

    excel = request.FILES.get("excel_file")
    if not excel:
        messages.error(request, "Seleccioná un archivo Excel.")
        return redirect(f"{reverse('lms:store_admin')}?tab=products")

    try:
        wb   = openpyxl.load_workbook(io.BytesIO(excel.read()), data_only=True)
        ws   = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        messages.error(request, f"No se pudo leer el archivo: {e}")
        return redirect(f"{reverse('lms:store_admin')}?tab=products")

    if not rows:
        messages.error(request, "El archivo está vacío.")
        return redirect(f"{reverse('lms:store_admin')}?tab=products")

    headers = [str(h).strip().lower() if h else "" for h in rows[0]]

    def col(name):
        try:
            return headers.index(name)
        except ValueError:
            return None

    created = updated = skipped = 0

    for i, row in enumerate(rows[1:], start=2):
        def get(name, default=""):
            idx = col(name)
            if idx is None:
                return default
            v = row[idx]
            return str(v).strip() if v is not None else default

        name = get("name")
        if not name:
            skipped += 1
            continue

        try:
            price = Decimal(get("price_ars", "0") or "0")
        except InvalidOperation:
            price = Decimal("0")

        try:
            stock = int(float(get("stock", "0") or "0"))
        except (ValueError, TypeError):
            stock = 0

        def to_decimal(field, default="0"):
            try:
                return Decimal(get(field, default) or default)
            except InvalidOperation:
                return Decimal(default)

        weight_kg   = to_decimal("weight_kg")
        width_cm    = to_decimal("width_cm")
        height_cm   = to_decimal("height_cm")
        depth_cm    = to_decimal("depth_cm")
        image_url   = get("image_url")
        description = get("description")
        active_raw  = get("is_active", "1").lower()
        is_active   = active_raw not in ("0", "false", "no", "")

        cat_name = get("category")
        category = None
        if cat_name:
            category, _ = ProductCategory.objects.get_or_create(
                name=cat_name,
                defaults={"slug": slugify(cat_name)}
            )

        existing = Product.objects.filter(name=name).first()
        if existing:
            existing.price_ars   = price
            existing.stock       = stock
            existing.description = description
            existing.weight_kg   = weight_kg
            existing.width_cm    = width_cm
            existing.height_cm   = height_cm
            existing.depth_cm    = depth_cm
            existing.image_url   = image_url
            existing.is_active   = is_active
            if category:
                existing.category = category
            existing.save()
            updated += 1
        else:
            base_slug = slugify(name) or f"producto-{i}"
            slug      = base_slug
            n         = 1
            while Product.objects.filter(slug=slug).exists():
                slug = f"{base_slug}-{n}"
                n   += 1
            Product.objects.create(
                name=name, slug=slug, price_ars=price, stock=stock,
                description=description, weight_kg=weight_kg,
                width_cm=width_cm, height_cm=height_cm, depth_cm=depth_cm,
                image_url=image_url, is_active=is_active, category=category,
            )
            created += 1

    messages.success(request, f"Importación completada: {created} creados, {updated} actualizados, {skipped} omitidos.")
    return redirect(f"{reverse('lms:store_admin')}?tab=products")


@user_passes_test(_is_staff)
def store_admin_product_import_template(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"
    ws.append(["name", "price_ars", "stock", "description",
               "weight_kg", "width_cm", "height_cm", "depth_cm",
               "image_url", "category", "is_active"])
    ws.append(["Ejemplo producto", 1500, 10, "Descripción del producto",
               0.5, 10, 5, 3, "https://...", "Categoría A", 1])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = 'attachment; filename="plantilla_productos.xlsx"'
    return resp


# ── CATEGORÍAS ──

@user_passes_test(_is_staff)
@require_POST
def store_admin_category_save(request):
    cat_id = request.POST.get("cat_id")
    name   = request.POST.get("name", "").strip()
    if not name:
        messages.error(request, "El nombre no puede estar vacío.")
        return redirect(f"{reverse('lms:store_admin')}?tab=categories")

    if cat_id:
        cat      = get_object_or_404(ProductCategory, id=cat_id)
        cat.name = name
        cat.save()
        messages.success(request, "Categoría actualizada.")
    else:
        ProductCategory.objects.create(name=name, slug=slugify(name))
        messages.success(request, "Categoría creada.")
    return redirect(f"{reverse('lms:store_admin')}?tab=categories")


@user_passes_test(_is_staff)
@require_POST
def store_admin_category_delete(request, cat_id):
    cat = get_object_or_404(ProductCategory, id=cat_id)
    cat.delete()
    messages.success(request, "Categoría eliminada.")
    return redirect(f"{reverse('lms:store_admin')}?tab=categories")


# ── PEDIDOS ──

@user_passes_test(_is_staff)
def store_admin_order_detail(request, order_id):
    order = get_object_or_404(StoreOrder, id=order_id)
    return render(request, "lms/store/admin_order_detail_tienda.html", {
        "order":    order,
        "statuses": StoreOrder.STATUS,
    })


@user_passes_test(_is_staff)
@require_POST
def store_admin_order_update(request, order_id):
    order           = get_object_or_404(StoreOrder, id=order_id)
    new_status      = request.POST.get("status")
    tracking_number = request.POST.get("tracking_number", "").strip()
    tracking_url    = request.POST.get("tracking_url", "").strip()
    notes           = request.POST.get("notes", "").strip()

    valid_statuses = [s[0] for s in StoreOrder.STATUS]
    if new_status in valid_statuses:
        order.status = new_status
    if tracking_number:
        order.tracking_number = tracking_number
    if tracking_url:
        order.tracking_url = tracking_url
    if notes:
        order.notes = notes
    order.save()
    messages.success(request, f"Pedido #{order.id} actualizado.")

    next_url = request.POST.get("next") or reverse("lms:store_admin_order_detail", args=[order.id])
    return redirect(next_url)


@user_passes_test(_is_staff)
@require_POST
def store_admin_order_delete(request, order_id):
    order = get_object_or_404(StoreOrder, id=order_id)
    oid   = order.id
    order.delete()
    messages.success(request, f"Pedido #{oid} eliminado.")
    return redirect(f"{reverse('lms:store_admin')}?tab=orders")


# ── ENVÍOS ──

@user_passes_test(_is_staff)
@require_POST
def store_admin_shipping_save(request):
    config                  = ShippingConfig.get()
    config.base_rate_ars    = Decimal(request.POST.get("base_rate_ars") or "0")
    config.per_kg_ars       = Decimal(request.POST.get("per_kg_ars") or "0")
    config.free_over_ars    = Decimal(request.POST.get("free_over_ars") or "0")
    config.provider_name    = request.POST.get("provider_name", "").strip()
    config.provider_url     = request.POST.get("provider_url", "").strip()
    config.provider_api_key = request.POST.get("provider_api_key", "").strip()
    config.provider_active  = request.POST.get("provider_active") == "on"
    config.save()
    messages.success(request, "Configuración de envíos guardada.")
    return redirect(f"{reverse('lms:store_admin')}?tab=shipping")


@user_passes_test(_is_staff)
@require_POST
def store_admin_zone_save(request):
    zone_id      = request.POST.get("zone_id")
    name         = request.POST.get("name", "").strip()
    multiplier   = Decimal(request.POST.get("multiplier") or "1")
    postal_codes = request.POST.get("postal_codes", "").strip()

    if zone_id:
        zone              = get_object_or_404(ShippingZone, id=zone_id)
        zone.name         = name
        zone.multiplier   = multiplier
        zone.postal_codes = postal_codes
        zone.save()
        messages.success(request, "Zona actualizada.")
    else:
        ShippingZone.objects.create(name=name, multiplier=multiplier, postal_codes=postal_codes)
        messages.success(request, "Zona creada.")
    return redirect(f"{reverse('lms:store_admin')}?tab=shipping")


@user_passes_test(_is_staff)
@require_POST
def store_admin_zone_delete(request, zone_id):
    zone = get_object_or_404(ShippingZone, id=zone_id)
    zone.delete()
    messages.success(request, "Zona eliminada.")
    return redirect(f"{reverse('lms:store_admin')}?tab=shipping")


# ── IMPORTACIÓN EXCEL — ZONAS ──

@user_passes_test(_is_staff)
@require_POST
def store_admin_zone_import(request):
    from decimal import InvalidOperation

    excel = request.FILES.get("excel_file")
    if not excel:
        messages.error(request, "Seleccioná un archivo Excel.")
        return redirect(f"{reverse('lms:store_admin')}?tab=shipping")

    try:
        wb   = openpyxl.load_workbook(io.BytesIO(excel.read()), data_only=True)
        ws   = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        messages.error(request, f"No se pudo leer el archivo: {e}")
        return redirect(f"{reverse('lms:store_admin')}?tab=shipping")

    if not rows:
        messages.error(request, "El archivo está vacío.")
        return redirect(f"{reverse('lms:store_admin')}?tab=shipping")

    headers = [str(h).strip().lower() if h else "" for h in rows[0]]

    def col(name):
        try:
            return headers.index(name)
        except ValueError:
            return None

    created = updated = skipped = 0

    for row in rows[1:]:
        def get(name, default=""):
            idx = col(name)
            if idx is None:
                return default
            v = row[idx]
            return str(v).strip() if v is not None else default

        name = get("name")
        if not name:
            skipped += 1
            continue

        try:
            multiplier = Decimal(get("multiplier", "1") or "1")
        except Exception:
            multiplier = Decimal("1")

        postal_codes = get("postal_codes")

        existing = ShippingZone.objects.filter(name=name).first()
        if existing:
            existing.multiplier   = multiplier
            existing.postal_codes = postal_codes
            existing.save()
            updated += 1
        else:
            ShippingZone.objects.create(name=name, multiplier=multiplier, postal_codes=postal_codes)
            created += 1

    messages.success(request, f"Zonas importadas: {created} creadas, {updated} actualizadas, {skipped} omitidas.")
    return redirect(f"{reverse('lms:store_admin')}?tab=shipping")


@user_passes_test(_is_staff)
def store_admin_zone_import_template(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Zonas"
    ws.append(["name", "multiplier", "postal_codes"])
    ws.append(["GBA", 1.5, "1600,1601,1602,1650"])
    ws.append(["Interior", 2.0, "5000,5001,3000"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = 'attachment; filename="plantilla_zonas.xlsx"'
    return resp


# ── NOTIFICACIONES DE PEDIDOS NUEVOS (polling) ────────────────────────────────

@user_passes_test(_is_staff)
@require_GET
def store_admin_new_orders(request):
    """
    Endpoint de polling para notificaciones.
    GET /panel-admin/tienda/pedidos/nuevos/?since=<unix_timestamp>
    Devuelve pedidos de tienda creados después de 'since'.
    """
    since_dt = None
    try:
        since_dt = datetime.fromtimestamp(int(request.GET.get("since", "")), tz=dt_timezone.utc)
    except (TypeError, ValueError):
        pass

    qs = StoreOrder.objects.select_related("user").order_by("-created_at")
    if since_dt:
        qs = qs.filter(created_at__gt=since_dt)

    orders = [
        {
            "id":         o.id,
            "recipient":  o.recipient_name,
            "total":      str(o.total_ars),
            "payment":    o.get_payment_method_display(),
            "created_at": o.created_at.isoformat(),
            "detail_url": reverse("lms:store_admin_order_detail", args=[o.id]),
        }
        for o in qs[:10]
    ]

    return JsonResponse({"count": len(orders), "orders": orders})